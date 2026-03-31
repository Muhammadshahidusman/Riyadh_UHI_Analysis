# ============================================================
# COMPLETE STATISTICAL ANALYSIS - LULC & UHI (Riyadh)
# Sections 5.6.1 to 5.6.4  — FULLY CORRECTED VERSION
# ============================================================

import os, math, warnings
import numpy as np
import pandas as pd
import geopandas as gpd
import rasterio
from rasterio.mask import mask
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.stats import pearsonr, norm
import libpysal.weights as lps_weights
from esda.getisord import G_Local
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

os.environ.update({
    "GDAL_CACHEMAX":      "4096",
    "GDAL_NUM_THREADS":   "ALL_CPUS",
    "CPL_CURL_GAZETTEER": "false",
    "LOKY_MAX_WORKERS":   "1",
})

# ============================================================
# 1. PATHS
# ============================================================
BASE_DIR     = r"C:\Users\SHEHER SAAZ\Music\Project\Urban_Heat\datasets\Riyadh\Statistical Analysis\Riyadh_Analysis"
DATASETS_DIR = os.path.join(BASE_DIR, "Datasets")

LULC_DIR     = os.path.join(DATASETS_DIR, "LUL_Classified")
UHI_DIR      = os.path.join(DATASETS_DIR, "UHI_Classified")
BOUNDARY_SHP = os.path.join(DATASETS_DIR, "Riyadh_Boundary", "Riyadh City.shp")
OUTPUT_EXCEL = os.path.join(BASE_DIR, "Riyadh_Statistical_Analysis_5_6.xlsx")
OUTPUT_MAPS  = os.path.join(BASE_DIR, "maps")
os.makedirs(OUTPUT_MAPS, exist_ok=True)

# ============================================================
# 2. FILE MAPPING
# ============================================================
YEARS = [1990, 1995, 2001, 2014, 2019, 2025]

LULC_FILES = {
    1990: os.path.join(LULC_DIR, "LULC_1990.tif"),
    1995: os.path.join(LULC_DIR, "LULC_1995.tif"),
    2001: os.path.join(LULC_DIR, "LULC_2001.tif"),
    2014: os.path.join(LULC_DIR, "LULC_2014.tif"),
    2019: os.path.join(LULC_DIR, "LULC_2019.tif"),
    2025: os.path.join(LULC_DIR, "LULC_2025.tif"),
}

UHI_FILES = {
    1990: os.path.join(UHI_DIR, "Reclass_UHI_1990.tif"),
    1995: os.path.join(UHI_DIR, "Reclass_UHI_1995.tif"),
    2001: os.path.join(UHI_DIR, "Reclass_UHI_2001.tif"),
    2014: os.path.join(UHI_DIR, "Reclass_UHI_2014.tif"),
    2019: os.path.join(UHI_DIR, "Reclass_UHI_2019.tif"),
    2025: os.path.join(UHI_DIR, "Reclass_UHI_2025.tif"),
}

# LULC: 0=Water, 1=BuiltUp, 2=Vegetation, 3=Barrenland
LULC_CLASSES = {
    0: "Water",
    1: "BuiltUp",
    2: "Vegetation",
    3: "Barrenland",
}

# UHI: 6 classes (values 1–6)
UHI_CLASSES = {
    1: "Extreme Cool Island",
    2: "Strong Cool Island",
    3: "Moderate Cool",
    4: "Moderate Hot",
    5: "Strong Hot",
    6: "Extreme Hot",
}

PIXEL_AREA_SQM = 900
SQM_TO_SQKM    = 1e-6

# ============================================================
# 3. HELPER FUNCTIONS
# ============================================================

def load_masked_raster(tif_path, boundary_gdf, zero_is_nodata=False):
    """
    zero_is_nodata=False → LULC (value 0 = Water, keep it)
    zero_is_nodata=True  → UHI  (value 0 = nodata, mask it)
    """
    with rasterio.open(tif_path) as src:
        boundary_reproj = boundary_gdf.to_crs(src.crs)
        geoms = [geom.__geo_interface__ for geom in boundary_reproj.geometry]
        out_img, out_transform = mask(src, geoms, crop=True)
        data = out_img[0].astype(np.uint16).astype(np.float32)
        if src.nodata is not None:
            data[data == float(src.nodata)] = np.nan
        data[data == 255] = np.nan
        if zero_is_nodata:
            data[data == 0] = np.nan
    return data, out_transform


def pixel_counts_to_area(arr, class_map):
    return {label: int(np.nansum(arr == val)) * PIXEL_AREA_SQM * SQM_TO_SQKM
            for val, label in class_map.items()}


def build_area_df(files_dict, class_map, boundary, zero_is_nodata=False):
    rows = []
    for year, fpath in files_dict.items():
        if not os.path.exists(fpath):
            print(f"  ⚠ Not found: {fpath}")
            continue
        arr, _ = load_masked_raster(fpath, boundary, zero_is_nodata=zero_is_nodata)
        area_dict = pixel_counts_to_area(arr, class_map)
        total = sum(area_dict.values())
        for cls, area in area_dict.items():
            rows.append({"Year": year, "Class": cls,
                         "Area_sq_km": round(area, 4),
                         "Percentage": round(area / total * 100, 4) if total > 0 else 0})
    return pd.DataFrame(rows)


def mann_kendall_sen(years, values):
    years  = np.array(years,  dtype=float)
    values = np.array(values, dtype=float)
    n = len(values)
    s, slopes = 0, []
    for i in range(n - 1):
        for j in range(i + 1, n):
            d = values[j] - values[i]
            s += np.sign(d)
            slopes.append(d / (years[j] - years[i]))
    var_s = n * (n - 1) * (2 * n + 5) / 18
    z   = (s - np.sign(s)) / math.sqrt(var_s) if s != 0 else 0.0
    p   = 2 * (1 - norm.cdf(abs(z)))
    tau = s / (n * (n - 1) / 2)
    sen = float(np.median(slopes))
    return {"S": s, "Tau": round(tau, 4), "Z": round(z, 4),
            "P_Value": round(p, 4),
            "Sen_Slope_sq_km_per_year": round(sen, 4),
            "Trend": "Increasing" if sen > 0 else ("Decreasing" if sen < 0 else "No trend"),
            "Significant_α0.05": "Yes" if p < 0.05 else "No"}


def transition_matrix(arr_t1, arr_t2, class_map):
    labels = list(class_map.keys())
    names  = [class_map[l] for l in labels]
    matrix = np.zeros((len(labels), len(labels)), dtype=np.int64)
    for i, fc in enumerate(labels):
        for j, tc in enumerate(labels):
            matrix[i, j] = int(np.nansum((arr_t1 == fc) & (arr_t2 == tc)))
    return pd.DataFrame(matrix * PIXEL_AREA_SQM * SQM_TO_SQKM,
                        index=[f"From: {n}" for n in names],
                        columns=[f"To: {n}" for n in names]).round(2)


def compute_gi_star(tif_path, boundary_gdf, sample_step=8):
    """
    FIXES APPLIED:
      1. sample_step 30 → 8  (denser sampling for spatial autocorrelation)
      2. Removed duplicate transform='r' from G_Local (was double-standardizing)
      3. zero_is_nodata=True for UHI rasters
    """
    arr, out_transform = load_masked_raster(tif_path, boundary_gdf, zero_is_nodata=True)
    rows_idx, cols_idx = np.where(~np.isnan(arr))
    rows_idx = rows_idx[::sample_step]
    cols_idx = cols_idx[::sample_step]
    values   = arr[rows_idx, cols_idx]

    with rasterio.open(tif_path) as src:
        crs = src.crs
    xs, ys = rasterio.transform.xy(out_transform, rows_idx, cols_idx)
    coords  = np.column_stack([xs, ys])

    w = lps_weights.KNN.from_array(coords, k=8)
    w.transform = 'r'                              # row-standardize ONCE only

    gi = G_Local(values, w, star=True, permutations=0)   # no transform= here
    p_values = 2 * (1 - norm.cdf(np.abs(gi.Zs)))

    def classify(z, p):
        if p >= 0.1:   return "Not Significant"
        if z >  2.58:  return "Hot Spot (99%)"
        if z >  1.96:  return "Hot Spot (95%)"
        if z >  1.65:  return "Hot Spot (90%)"
        if z < -2.58:  return "Cold Spot (99%)"
        if z < -1.96:  return "Cold Spot (95%)"
        if z < -1.65:  return "Cold Spot (90%)"
        return "Not Significant"

    gdf = gpd.GeoDataFrame({
        "x": xs, "y": ys, "UHI_Value": values,
        "Gi_Z_Score":    gi.Zs,
        "Gi_P_Value":    p_values,
        "Hotspot_Class": [classify(z, p) for z, p in zip(gi.Zs, p_values)]
    }, geometry=gpd.points_from_xy(xs, ys), crs=crs)
    return gdf


# ============================================================
# 4. EXCEL HELPERS
# ============================================================
def style_sheet(ws, header_color="1F4E79"):
    hfill  = PatternFill("solid", fgColor=header_color)
    hfont  = Font(bold=True, color="FFFFFF", size=10)
    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

def write_df(wb, sheet_name, df, header_color="1F4E79"):
    ws = wb.create_sheet(sheet_name[:31])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    style_sheet(ws, header_color)
    return ws

# ============================================================
# 5. MAIN PIPELINE
# ============================================================
print("="*65)
print("  RIYADH LULC & UHI - STATISTICAL ANALYSIS  5.6.1 - 5.6.4")
print("="*65)

boundary = gpd.read_file(BOUNDARY_SHP)

# ── Build area tables ────────────────────────────────────────
print("\n[1/7] Building LULC area table...")
lulc_df = build_area_df(LULC_FILES, LULC_CLASSES, boundary, zero_is_nodata=False)
print("[2/7] Building UHI area table...")
uhi_df  = build_area_df(UHI_FILES,  UHI_CLASSES,  boundary, zero_is_nodata=True)

lulc_wide = lulc_df.pivot(index="Class", columns="Year", values="Area_sq_km").fillna(0)
uhi_wide  = uhi_df.pivot( index="Class", columns="Year", values="Area_sq_km").fillna(0)

# ── 5.6.1 CHANGE DETECTION ──────────────────────────────────
print("[3/7] 5.6.1 Change Detection & Areal Transition...")

def change_table(wide):
    avail = sorted(c for c in wide.columns if isinstance(c, int))
    ct = wide[avail].copy()
    for i in range(len(avail) - 1):
        ct[f"Chg_{avail[i]}_{avail[i+1]}"] = wide[avail[i+1]] - wide[avail[i]]
    ct["Net_Chg_Total"]     = wide[avail[-1]] - wide[avail[0]]
    ct["Pct_Chg_Total_%"]   = np.where(wide[avail[0]] != 0,
        (wide[avail[-1]] - wide[avail[0]]) / wide[avail[0]] * 100, np.nan)
    ct["Ann_Rate_sq_km_yr"] = (wide[avail[-1]] - wide[avail[0]]) / (avail[-1] - avail[0])
    return ct.round(4)

lulc_change = change_table(lulc_wide)
uhi_change  = change_table(uhi_wide)

transition_dfs = {}
years_ok = sorted(y for y in YEARS if os.path.exists(LULC_FILES.get(y, "")))
for i in range(len(years_ok) - 1):
    y1, y2 = years_ok[i], years_ok[i+1]
    a1, _  = load_masked_raster(LULC_FILES[y1], boundary, zero_is_nodata=False)
    a2, _  = load_masked_raster(LULC_FILES[y2], boundary, zero_is_nodata=False)
    r = min(a1.shape[0], a2.shape[0]); c = min(a1.shape[1], a2.shape[1])
    transition_dfs[f"Trans_{y1}_{y2}"] = transition_matrix(a1[:r,:c], a2[:r,:c], LULC_CLASSES)

# ── 5.6.2 MANN-KENDALL + SEN ────────────────────────────────
print("[4/7] 5.6.2 Mann-Kendall & Sen's Slope...")

def mk_table(wide, label_col):
    avail = sorted(c for c in wide.columns if isinstance(c, int))
    rows  = []
    for cls in wide.index:
        res = mann_kendall_sen(avail, wide.loc[cls, avail].values.astype(float))
        res[label_col] = cls
        rows.append(res)
    cols = [label_col, "S", "Tau", "Z", "P_Value",
            "Sen_Slope_sq_km_per_year", "Trend", "Significant_α0.05"]
    return pd.DataFrame(rows)[cols]

mk_lulc = mk_table(lulc_wide, "LULC_Class")
mk_uhi  = mk_table(uhi_wide,  "UHI_Zone")

# ── 5.6.3 PEARSON + REGRESSION ──────────────────────────────
print("[5/7] 5.6.3 Pearson Correlation & Multiple Regression...")

avail_years = sorted(set(lulc_wide.columns) & set(uhi_wide.columns)
                     & {c for c in lulc_wide.columns if isinstance(c, int)})

corr_rows = []
for lc in lulc_wide.index:
    x = lulc_wide.loc[lc, avail_years].values.astype(float)
    for zone in uhi_wide.index:
        y = uhi_wide.loc[zone, avail_years].values.astype(float)
        r, p = (pearsonr(x, y) if np.std(x) > 0 and np.std(y) > 0
                else (np.nan, np.nan))
        corr_rows.append({
            "LULC_Class":        lc,
            "UHI_Zone":          zone,
            "Pearson_r":         round(r,    4) if pd.notna(r) else np.nan,
            "P_Value":           round(p,    4) if pd.notna(p) else np.nan,
            "R_Squared":         round(r**2, 4) if pd.notna(r) else np.nan,
            "Significant_α0.05": "Yes" if pd.notna(p) and p < 0.05 else "No",
            "Direction": ("Positive" if pd.notna(r) and r > 0 else
                          "Negative" if pd.notna(r) and r < 0 else "NA")
        })
pearson_table = pd.DataFrame(corr_rows).sort_values("Pearson_r", key=abs, ascending=False)

X = lulc_wide[avail_years].T.values
reg_rows = []
for zone in uhi_wide.index:
    y    = uhi_wide.loc[zone, avail_years].values.astype(float)
    n, k = len(avail_years), X.shape[1]
    if n <= k:
        row = {"UHI_Zone": zone,
               "R_Squared": "n<=k - see Pearson",
               "Note": f"Only {n} observations for {k} predictors. Use Pearson instead."}
    else:
        X_b = np.column_stack([np.ones(n), X])
        coeffs, _, _, _ = np.linalg.lstsq(X_b, y, rcond=None)
        yp     = X_b @ coeffs
        ss_res = np.sum((y - yp)**2)
        ss_tot = np.sum((y - np.mean(y))**2)
        r2     = 1 - ss_res / ss_tot if ss_tot > 0 else np.nan
        row = {"UHI_Zone": zone, "R_Squared": round(r2, 4),
               "Intercept": round(coeffs[0], 4)}
        for ci, c in enumerate(lulc_wide.index):
            row[f"Coeff_{c}"] = round(coeffs[ci+1], 4)
        row["Note"] = "OLS Multiple Regression"
    reg_rows.append(row)
regression_table = pd.DataFrame(reg_rows)

# ── 5.6.4 GETIS-ORD Gi* ─────────────────────────────────────
print("[6/7] 5.6.4 Getis-Ord Gi* Spatial Hotspot Analysis...")

HOTSPOT_COLORS = {
    "Hot Spot (99%)":  "#d73027",
    "Hot Spot (95%)":  "#f46d43",
    "Hot Spot (90%)":  "#fdae61",
    "Not Significant": "#ffffbf",
    "Cold Spot (90%)": "#abd9e9",
    "Cold Spot (95%)": "#74add1",
    "Cold Spot (99%)": "#4575b4",
}

gi_summary_rows = []
for year, fpath in UHI_FILES.items():
    if not os.path.exists(fpath):
        print(f"  ⚠ Skipping {year} - file not found"); continue
    print(f"  Gi* -> {year}...")
    gdf = compute_gi_star(fpath, boundary, sample_step=8)
    for cls, grp in gdf.groupby("Hotspot_Class"):
        gi_summary_rows.append({
            "Year":          year,
            "Hotspot_Class": cls,
            "Point_Count":   len(grp),
            "Pct_of_Total":  round(len(grp) / len(gdf) * 100, 2),
            "Mean_Gi_Z":     round(grp["Gi_Z_Score"].mean(), 4)
        })
    fig, ax = plt.subplots(figsize=(10, 8))
    boundary.to_crs(gdf.crs).boundary.plot(ax=ax, color="black", linewidth=1.5)
    for cls, color in HOTSPOT_COLORS.items():
        sub = gdf[gdf["Hotspot_Class"] == cls]
        if len(sub):
            sub.plot(ax=ax, markersize=1.5, color=color, label=cls, alpha=0.7)
    ax.set_title(f"Getis-Ord Gi* - UHI Hotspot Analysis  ({year})",
                 fontsize=14, fontweight='bold')
    ax.legend(loc="lower right", fontsize=7, title="Hotspot Class",
              markerscale=5, framealpha=0.9)
    ax.set_axis_off(); plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_MAPS, f"GiStar_UHI_{year}.png"),
                dpi=200, bbox_inches='tight')
    plt.close()

gi_summary = pd.DataFrame(gi_summary_rows)

# ── WRITE EXCEL ──────────────────────────────────────────────
print("[7/7] Writing Excel workbook...")

wb = openpyxl.Workbook()
wb.remove(wb.active)

readme = pd.DataFrame({
    "Section":  ["5.6.1","5.6.1","5.6.2","5.6.2","5.6.3","5.6.3","5.6.4"],
    "Analysis": ["Change Detection","Areal Transition","Mann-Kendall","Sen Slope",
                 "Pearson Correlation","Multiple Regression","Getis-Ord Gi*"],
    "Sheet":    ["LULC/UHI_Change_Detection","Trans_YYYY_YYYY (one per pair)",
                 "MK_Sen_LULC / MK_Sen_UHI","MK_Sen_LULC / MK_Sen_UHI (Sen_Slope column)",
                 "Pearson_Correlation","Multiple_Regression","GiStar_Summary + PNG maps"]
})
write_df(wb, "Read_Me",               readme,                    "1F3864")
write_df(wb, "LULC_Area_Long",        lulc_df,                   "1F3864")
write_df(wb, "UHI_Area_Long",         uhi_df,                    "1F3864")
write_df(wb, "LULC_Change_Detection", lulc_change.reset_index(), "C55A11")
write_df(wb, "UHI_Change_Detection",  uhi_change.reset_index(),  "C55A11")
for key, tdf in transition_dfs.items():
    write_df(wb, key, tdf.reset_index(), "833C11")
write_df(wb, "MK_Sen_LULC",          mk_lulc,                   "375623")
write_df(wb, "MK_Sen_UHI",           mk_uhi,                    "375623")
write_df(wb, "Pearson_Correlation",   pearson_table,             "4472C4")
write_df(wb, "Multiple_Regression",   regression_table,          "4472C4")
write_df(wb, "GiStar_Summary",        gi_summary,                "7030A0")

wb.save(OUTPUT_EXCEL)
print(f"\n[DONE] Excel  -> {OUTPUT_EXCEL}")
print(f"[DONE] Maps   -> {OUTPUT_MAPS}")